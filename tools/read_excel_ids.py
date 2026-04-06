import openpyxl
import sys
import os

excel_path = os.path.join(os.path.dirname(__file__), '..', 'Tavlor dokumentation Fredrik Beckman.xlsx')
excel_path = os.path.abspath(excel_path)
print(f"Reading: {excel_path}")

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

# Print headers (row 2)
headers = [ws.cell(row=2, column=c).value for c in range(1, 10)]
print("Headers:", headers)

target_ids = {269, 270, 271}

for row in range(3, 600):
    namn = ws.cell(row=row, column=1).value
    if namn is None:
        continue
    namn = str(namn).strip()
    import re
    m = re.match(r'^(\d+)', namn)
    if m:
        id_num = int(m.group(1))
        if id_num in target_ids:
            storlek  = ws.cell(row=row, column=2).value
            anmarkning = ws.cell(row=row, column=4).value
            tillverkad = ws.cell(row=row, column=5).value
            material   = ws.cell(row=row, column=6).value
            print(f"\nID: {id_num}")
            print(f"  Namn: {namn}")
            print(f"  Storlek: {storlek}")
            print(f"  Tillverkad: {tillverkad}")
            print(f"  Material: {material}")
            print(f"  Anmärkning: {anmarkning}")
