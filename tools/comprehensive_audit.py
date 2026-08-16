import os
import re
import json
import openpyxl

print("=== 1. AUDITING EXCEL FILE ===")
wb = openpyxl.load_workbook('Tavlor dokumentation Fredrik Beckman.xlsx', data_only=True)
print("Sheet names:", wb.sheetnames)
sheet = wb.active
rows = list(sheet.iter_rows(values_only=True))
print(f"Total rows in active sheet: {len(rows)}")
header = rows[0]
print("Header:", header)

excel_paintings = []
for i, r in enumerate(rows[1:], start=2):
    if not any(r):
        continue
    # Let's see how columns are mapped
    excel_paintings.append({
        'row_idx': i,
        'data': r
    })
print(f"Total non-empty data rows in Excel: {len(excel_paintings)}")
if excel_paintings:
    print("Sample row 1:", excel_paintings[0])
    print("Sample last row:", excel_paintings[-1])

print("\n=== 2. AUDITING DATA.JS ===")
with open('js/data.js', 'r', encoding='utf-8', errors='replace') as f:
    data_content = f.read()

# Extract the JSON array from GALLERY_IMAGES = [...]
match = re.search(r'const\s+GALLERY_IMAGES\s*=\s*(\[[\s\S]*?\]);', data_content)
if match:
    # Need to be careful with trailing commas in JS
    js_arr_str = match.group(1)
    # Remove trailing commas before closing braces/brackets
    clean_js = re.sub(r',\s*([\]}])', r'\1', js_arr_str)
    try:
        gallery_images = json.loads(clean_js)
        print(f"Successfully parsed {len(gallery_images)} items from js/data.js")
    except Exception as e:
        print("JSON parse error, falling back to regex:", e)
        gallery_images = []
else:
    print("Could not find GALLERY_IMAGES regex match")
    gallery_images = []

print("\n=== 3. AUDITING ASSETS/IMAGES ===")
asset_files = set(os.listdir('assets/images')) if os.path.exists('assets/images') else set()
print(f"Total files in assets/images: {len(asset_files)}")

print("\n=== 4. CHECKING ID CONTINUITY AND DUPLICATES IN DATA.JS ===")
ids = [item.get('id') for item in gallery_images if 'id' in item]
from collections import Counter
id_counts = Counter(ids)
duplicates = {k: v for k, v in id_counts.items() if v > 1}
print(f"Duplicate IDs in data.js: {duplicates}")

all_ids = set(ids)
max_id = max(ids) if ids else 0
print(f"Min ID: {min(ids) if ids else 0}, Max ID: {max_id}, Unique IDs count: {len(all_ids)}")
missing_ids = [i for i in range(1, max_id + 1) if i not in all_ids]
print(f"Missing IDs between 1 and {max_id}: {missing_ids}")

print("\n=== 5. CHECKING MISSING FILES (DATA.JS -> ASSETS/IMAGES) ===")
missing_files = []
for item in gallery_images:
    fn = item.get('filename')
    if fn not in asset_files:
        missing_files.append((item.get('id'), item.get('title'), fn))

print(f"Missing files referenced in data.js ({len(missing_files)}):")
for item_id, title, fn in missing_files:
    print(f"  ID {item_id}: '{title}' -> filename '{fn}'")

print("\n=== 6. CHECKING FILES ON DISK NOT IN DATA.JS ===")
used_filenames = set(item.get('filename') for item in gallery_images)
unused_files = [f for f in asset_files if f not in used_filenames]
print(f"Files in assets/images NOT referenced in data.js ({len(unused_files)}):")
for f in sorted(unused_files):
    print(f"  Disk file: '{f}'")

print("\n=== 7. CHECKING PROBLEMATIC CHARACTERS IN FILENAMES ===")
problematic = []
for f in asset_files:
    if '#' in f or '%' in f or '?' in f or '&' in f or '^' in f or '`' in f:
        problematic.append(f)
print(f"Problematic filenames in assets/images ({len(problematic)}):")
for f in problematic:
    print(f"  '{f}'")
